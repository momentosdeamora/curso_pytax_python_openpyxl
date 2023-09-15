from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('plan_teste.xlsx')
ws = wb['teste1']


cell_range = ws['A1:C4']
print(cell_range)
for x, y, z in cell_range:
    print(x.value, y.value, z.value)

#list comprehention

dados = [(x.value, y.value, z.value) for x, y, z in cell_range]
print(dados)

range_cell = ws[get_column_letter(1) + '1', get_column_letter(3) + '4']
for x, y, z in range_cell:
    print(x.value, y.valuie, z.value)
print(range_cell)