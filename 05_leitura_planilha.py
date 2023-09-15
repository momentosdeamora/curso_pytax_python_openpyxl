from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

wb = load_workbook("primeira_planilha.xlsx")
wb.active

teste = wb['teste']

teste['A1'] = 'TESTE'

teste.cell(row=2,column=2,value=1000)

for x in range(3,20):
    for y in range(1,20):
        teste.cell(row=x, column=y, value=10)
wb.save('primeira_planilha.xlsx')