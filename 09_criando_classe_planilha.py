from openpyxl import Workbook, load_workbook

class Planilhas():
    def __init__(self, wb):
        self.wb = wb

    def save_wb(self, filename):
        self.wb.save(f"{filename}.xlsx")
    
    def open_wb(self, filename):
        self.wb = load_workbook(filename)

    def create_sheets(self, sheets):
        for s in sheets:
            self.wb.create_sheet(s)

    def delete_sheet_by_name(self, sheetName):
        self.wb.remove(self.wb[sheetName])

    def copy_sheet(self, sheetName, new_sheet=''):
        if new_sheet != "":
            ws = self.wb[sheetName]
            n_sheet = self.wb.copy_worksheet(ws)
            n_sheet.title = new_sheet

    def get_sheet_by_index(self, index):
        try:
            print(self.wb.worksheets[index])
            return self.wb.worksheets[index]
        except IndexError:
            print("Planilha não existe", index)

    def fill_data(self, sheet, lista):
        ws = self.wb[sheet]
        for x, data in enumerate(lista):
            for y, item in enumerate(data):
                ws.cell(row=x+1, column=y+1, value=item)

my_plan = Planilhas(Workbook())
my_plan.create_sheets(['teste1', 'teste2', 'teste3'])
lista = [
    ['CODIGO', 'DESCRICAO', 'VALOR'],
    ['001', 'MATERIAL DE LIMPEZA', '1000'],
    ['002', 'MATERIAL DE CONSUMO', '2000'],
    ['003', 'MATERIA PRIMA', '5000'],
]
my_plan.fill_data('teste1', lista)
my_plan.save_wb('plan_teste')