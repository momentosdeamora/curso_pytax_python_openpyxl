from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('plan_teste.xlsx')
ws = wb['teste1']

print(ws.max_row, ws.max_column)
print(ws.min_row, ws.min_column)

for row in ws.iter_rows(min_row=ws.min_row, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        print(cell.value)

for i in range(1, ws.max_row+1):
    print(ws['A' + str(i)].value)

print(ws.max_row)