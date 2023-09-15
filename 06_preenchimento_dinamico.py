from openpyxl import Workbook, load_workbook

wb = Workbook()

lista = [
    ['CODIGO', 'DESCRICAO', 'VALOR'],
    ['001', 'MATERIAL DE LIMPEZA', '1000'],
    ['002', 'MATERIAL DE CONSUMO', '2000'],
    ['003', 'MATERIA PRIMA', '5000'],
]

mw = wb['Minha planilha']

for x, data in enumerate(lista):
    for y, item in enumerate(data):
        mw.cell(row=x+1, column=y+1, value=item)

wb.save('primeira_planilha.xlsx')