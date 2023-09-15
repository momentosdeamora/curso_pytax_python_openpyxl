from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('Exemplo_2.xlsx')
ws = wb.active

def view_sheet(ws):
    intervalo = ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column)
    intervalo = [[x.value for x in row] for row in intervalo]
    df = pd.DataFrame(intervalo)
    df.rename(columns=df.iloc[0], inplace=True)
    df.drop([0], inplace=True)
    return df

#Inserir uma linha
ws.insert_rows(1,4)
view_sheet(ws)

#Remover linhas
ws.delete_rows(1,4)
view_sheet(ws)

#Remover colunas
ws.delete_cols(4,1)
view_sheet(ws)

#Inserir colunas
ws.insert_cols(2,5)
view_sheet(ws)
