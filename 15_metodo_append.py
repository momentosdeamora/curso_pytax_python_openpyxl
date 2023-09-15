from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('Exemplo_3.xlsx')
ws = wb.active

data = ['11122233000144', '2022-03-31', '5856', 'IRRF', '170.798,59']
ws.apprend(data)

data = ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column)
data = [[x.value for x in row] for row in data]

for row in data:
    ws.append(row)
