from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('teste', 0)
print(wb.sheetnames)