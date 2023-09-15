from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('Desafio_PyTax.xlsx')

dados = []
campos = ['COD IMPOSTO', 'COMPETENCIA', 'CNPJ', 'NOME', 'VALOR INSS', 'VALOR OUTRAS', 'VALOR TOTAL']
dados.append(campos)

ws = wb.active
#ws = wb["Planilha1"]

for column in ws.iter_cols(max_col=ws.max_column, max_row=1):
    for cell in column:
        if cell.value == "COD":
            cod = cell.offset(0,1).value
            competencia = cell.offset(1,1).value
            cnpj = cell.offset(2,1).value
            nome = cell.offset(5,0).value
            inss = cell.offset(3,3).value
            outras = cell.offset(8,3).value
            valor_total = cell.offset(11,3).value

            dados.append([cod, competencia, cnpj, nome, inss, outras, valor_total])

print(pd.DataFrame(dados))