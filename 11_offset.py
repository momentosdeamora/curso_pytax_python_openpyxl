from openpyxl import Workbook, load_workbook

wb = load_workbook("plan_teste.xlsx")
ws = wb['teste1']

print(ws.cell(1,2).value)
print(ws.cell(1,2).offset(3,1).value)

referencia = ws.cell(1,1)
referencia.offset(0,3).value = "OBSERVAÇÃO"
referencia.offset(1,3).value = "valor acima do esperado"
wb.save("plan_teste.xlsx")