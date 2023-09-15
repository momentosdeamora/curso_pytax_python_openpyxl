from openpyxl import Workbook, load_workbook

wb = load_workbook('segunda_planilha.xlsx')

ws = wb["Minha planilha"]

for row in ws.values:
    for value in row:
        print(value)

for row in ws.values:
    print(row)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=4, values_only=True):
    print(row)