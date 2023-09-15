from openpyxl import Workbook, load_workbook

wb = load_workbook("segunda_planilha.xlsx")

mp = wb['Minha planilha']
mp_copy = wb.copy_worksheet(mp)
mp_copy.title = "Copy"

wb.sheetnames

wb.save('teste.xlsx')

def copiar_planilha(nome, nome_nova_planilha):
    wb = load_workbook("teste.xlsx")
    ws = wb[nome]
    copia = wb.copy_worksheet(ws)
    copia.title = nome_nova_planilha

    wb.save('teste.xlsx')

copiar_planilha('Minha planilha', 'nova_copia')